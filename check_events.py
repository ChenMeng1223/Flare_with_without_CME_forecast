#!/usr/bin/env python3
"""检查 events.xlsx 的结构和数据"""
import openpyxl

wb = openpyxl.load_workbook('data/raw/events.xlsx')
ws = wb.active

# 获取列名
headers = [cell.value for cell in ws[1]]
print("列名：")
for i, h in enumerate(headers, 1):
    print(f"  {i:2d}. {h}")

# 获取前3行数据
print("\n前3行数据：")
for row_idx in range(2, min(5, ws.max_row + 1)):
    row_data = [cell.value for cell in ws[row_idx]]
    print(f"\n行 {row_idx}:")
    for i, (header, value) in enumerate(zip(headers, row_data)):
        if header in ['CME_associated', 'flare_class', 'event_id', 'start_time']:
            print(f"  {header}: {value} (类型: {type(value).__name__})")

# 查看 CME_associated 列的唯一值
print("\n\nCME_associated 列的唯一值：")
cme_col_idx = headers.index('CME_associated') + 1
cme_values = set()
for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=cme_col_idx).value
    cme_values.add(val)
print(sorted(cme_values))
